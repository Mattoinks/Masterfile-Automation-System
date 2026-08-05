"""Writes DN-breakdown rows (columns A-H) to the '2526' worksheet.

Independent from app/services/excel_layout.py / excel_service.py: those are
built for FY2526's one-row-per-record layout driven by field_mapping.json.
The 2526 sheet's later columns (I onward) are a separate, manually-driven
workflow (Phase 2) and are never touched here.
"""

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from config.settings import resolve_masterfile_path

WORKSHEET_NAME = "2526"
DATE_NUMBER_FORMAT = r"[$-4809]d\ mmm\ yyyy;@"
# The sheet's template only pre-formats a handful of rows with centered
# alignment; rows written beyond that range fall back to Excel's default
# (left/bottom) unless we set this explicitly on every cell we write.
_CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# Column A-H, in order.
_COLUMNS = [
    "test_bau",
    "original_label_lot_no",
    "date_code",
    "return_qty_from_dc",
    "disposition_or_ss_plan_name",
    "date_attached_ss_plan",
    "lw",
]

# Columns I-M. Column J (created_lot_no) is deterministic from that same
# row's original_label_lot_no and is written at intake time, on the same
# row as columns B-H. Columns I, K, L, M can only be known once the
# physical split actually happens - filled in later, on that same row,
# via update_row_lot_creation_fields() (or, for a row that needs to be
# added beyond the 1:1 original-lot mapping, via the secondary
# append_lot_creation_entry() insert path).
_LOT_CREATION_COLUMN_OFFSET = 9  # column I
_LOT_CREATION_FIELDS = [
    "date_created",
    "created_lot_no",
    "created_date_code",
    "physical_lot_qty",
    "lot_code",
]


class Lot2526ExcelError(Exception):
    pass


class Lot2526ExcelLockedError(Lot2526ExcelError):
    pass


class Lot2526ExcelWriter:
    def __init__(self, worksheet_name: str = WORKSHEET_NAME) -> None:
        self.worksheet_name = worksheet_name

    def _masterfile_path(self) -> Path:
        return resolve_masterfile_path()

    def _open_workbook(self):
        path = self._masterfile_path()
        if not path.exists():
            raise Lot2526ExcelError(f"Masterfile not found at {path}")
        try:
            return load_workbook(path)
        except PermissionError as exc:
            raise Lot2526ExcelLockedError(
                "Master Excel file is locked. Close it in Excel and try again."
            ) from exc

    def _get_sheet(self, wb):
        if self.worksheet_name not in wb.sheetnames:
            raise Lot2526ExcelError(
                f"Worksheet '{self.worksheet_name}' not found in masterfile"
            )
        return wb[self.worksheet_name]

    @staticmethod
    def _max_case_no(ws) -> int:
        max_no = 0
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=1).value
            if value is None:
                continue
            try:
                max_no = max(max_no, int(value))
            except (TypeError, ValueError):
                continue
        return max_no

    @staticmethod
    def _next_blank_row(ws) -> int:
        """One row past the *last* populated row in the sheet (columns
        A-H). A single blank separator row can sit between two existing
        cases - stopping at the first all-blank row found while scanning
        from the top would land there and overwrite the case that follows
        it, so this scans the whole sheet for the true last occupied row
        instead of stopping early."""
        last_col = 1 + len(_COLUMNS)  # A (case no) + the 7 breakdown columns
        last_populated = 1  # header row
        for row in range(2, ws.max_row + 1):
            if any(ws.cell(row=row, column=c).value is not None for c in range(1, last_col + 1)):
                last_populated = row
        return last_populated + 1

    def get_next_case_no(self) -> int:
        wb = self._open_workbook()
        try:
            ws = self._get_sheet(wb)
            return self._max_case_no(ws) + 1
        finally:
            wb.close()

    @staticmethod
    def _group_by_filename(fields_list: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
        """Consecutive entries sharing the same filename (or record_id, for
        callers that don't set filename) belong to one case. An entry with
        neither becomes its own single-row group - safe fallback, never
        merges unrelated rows."""
        groups: list[list[dict[str, Any]]] = []
        current_key: Any = object()
        current_group: list[dict[str, Any]] = []
        for fields in fields_list:
            key = fields.get("filename") or fields.get("record_id") or object()
            if current_group and key != current_key:
                groups.append(current_group)
                current_group = []
            current_group.append(fields)
            current_key = key
        if current_group:
            groups.append(current_group)
        return groups

    @staticmethod
    def _write_case_group(ws, start_row: int, case_no: int, group: list[dict[str, Any]]) -> None:
        created_lot_col = _LOT_CREATION_COLUMN_OFFSET + 1  # column J
        for i, fields in enumerate(group):
            row = start_row + i
            if i == 0:
                case_cell = ws.cell(row=row, column=1, value=case_no)
                case_cell.alignment = _CELL_ALIGNMENT
            for offset, field in enumerate(_COLUMNS, start=2):
                if field == "test_bau" and i > 0:
                    continue  # only the case's first row shows Test Bau
                value = fields.get(field)
                if value in (None, ""):
                    continue
                cell = ws.cell(row=row, column=offset, value=value)
                cell.alignment = _CELL_ALIGNMENT
                if field == "date_attached_ss_plan":
                    cell.number_format = DATE_NUMBER_FORMAT
            # Created Lot# (column J) is per-lot-line, unlike Test Bau -
            # every row gets its own, computed once at intake.
            created_lot_no = fields.get("created_lot_no")
            if created_lot_no not in (None, ""):
                cell = ws.cell(row=row, column=created_lot_col, value=created_lot_no)
                cell.alignment = _CELL_ALIGNMENT

    @staticmethod
    def _write_subtotal_row(ws, row: int, group: list[dict[str, Any]]) -> None:
        total = 0.0
        for fields in group:
            try:
                total += float(str(fields.get("return_qty_from_dc") or "0").replace(",", ""))
            except (TypeError, ValueError):
                continue
        total_value: float | int = int(total) if total == int(total) else total
        col = 2 + _COLUMNS.index("return_qty_from_dc")
        cell = ws.cell(row=row, column=col, value=total_value)
        cell.font = Font(bold=True)
        cell.alignment = _CELL_ALIGNMENT

    def append_breakdown_rows(self, fields_list: list[dict[str, Any]]) -> list[int]:
        """Appends one case per group of consecutive same-DN entries in
        fields_list (grouped by filename/record_id), opening and saving the
        workbook exactly once for the whole batch. Within a case: the first
        row gets No + Test Bau, remaining lot lines leave those blank; a
        bold subtotal row (sum of Return Qty from DC) follows when a case
        has more than one lot line, then a blank separator row before the
        next case. Leaves columns I onward untouched. Returns the case
        numbers written, one per case/group.
        """
        if not fields_list:
            return []
        wb = self._open_workbook()
        try:
            ws = self._get_sheet(wb)
            row = self._next_blank_row(ws)
            case_no = self._max_case_no(ws) + 1
            case_numbers: list[int] = []

            for group in self._group_by_filename(fields_list):
                self._write_case_group(ws, row, case_no, group)
                case_numbers.append(case_no)
                row += len(group)
                if len(group) > 1:
                    self._write_subtotal_row(ws, row, group)
                    row += 1
                row += 1  # blank separator row before the next case
                case_no += 1

            wb.save(self._masterfile_path())
            return case_numbers
        except PermissionError as exc:
            raise Lot2526ExcelLockedError(
                "Master Excel file is locked. Close it in Excel and try again."
            ) from exc
        finally:
            wb.close()

    def append_breakdown_row(self, fields: dict[str, Any]) -> int:
        """Appends one single-lot-line case (columns A-H only) to the 2526
        sheet. Leaves columns I onward untouched. Returns the case number
        written.
        """
        return self.append_breakdown_rows([fields])[0]

    @staticmethod
    def _is_subtotal_row(ws, row: int) -> bool:
        """A/B/C blank with a value in E (Return Qty from DC) - matches
        exactly what _write_subtotal_row produces, nothing else does."""
        col_a = ws.cell(row=row, column=1).value
        col_b = ws.cell(row=row, column=2).value
        col_c = ws.cell(row=row, column=3).value
        col_e = ws.cell(row=row, column=5).value
        return col_a is None and col_b is None and col_c is None and col_e is not None

    @staticmethod
    def _find_case_span(ws, case_no: int) -> tuple[int, int] | None:
        """Returns (start_row, end_row) inclusive for the given case
        number - end_row is the last non-blank row (across A-H and I-M)
        before the next fully-blank separator row or the next case's start.
        Always recomputed fresh; never cached, since row positions shift
        whenever anything is inserted elsewhere in the sheet."""
        start_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == case_no:
                start_row = row
                break
        if start_row is None:
            return None

        last_lot_col = _LOT_CREATION_COLUMN_OFFSET + len(_LOT_CREATION_FIELDS) - 1
        end_row = start_row
        row = start_row + 1
        while row <= ws.max_row:
            case_val = ws.cell(row=row, column=1).value
            if case_val is not None and case_val != case_no:
                break
            if not any(ws.cell(row=row, column=c).value is not None for c in range(1, last_lot_col + 1)):
                break
            end_row = row
            row += 1
        return start_row, end_row

    @staticmethod
    def _format_cell_date(value: Any) -> str:
        if value is None:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
        return str(value)

    def list_cases(self) -> list[dict[str, Any]]:
        """Scans the sheet fresh (no index/database) and returns a summary
        per case: case_no, test_bau, how many original lot lines it has,
        how many have their physical lot-creation recorded (Physical Lot
        Qty filled - the strongest signal an engineer has actually visited
        that row after the split happened), and the total Return Qty from
        DC."""
        wb = self._open_workbook()
        try:
            ws = self._get_sheet(wb)
            cases: list[dict[str, Any]] = []
            row = 2
            physical_qty_col = _LOT_CREATION_COLUMN_OFFSET + 3
            while row <= ws.max_row:
                case_no = ws.cell(row=row, column=1).value
                if not isinstance(case_no, (int, float)):
                    row += 1
                    continue
                case_no = int(case_no)
                span = self._find_case_span(ws, case_no)
                if span is None:
                    row += 1
                    continue
                start_row, end_row = span
                test_bau = ws.cell(row=start_row, column=2).value or ""
                lot_line_count = 0
                lot_creation_count = 0
                total_qty = 0.0
                for r in range(start_row, end_row + 1):
                    if self._is_subtotal_row(ws, r):
                        continue
                    if ws.cell(row=r, column=3).value not in (None, ""):
                        lot_line_count += 1
                        try:
                            total_qty += float(str(ws.cell(row=r, column=5).value or 0).replace(",", ""))
                        except (TypeError, ValueError):
                            pass
                    if ws.cell(row=r, column=physical_qty_col).value not in (None, ""):
                        lot_creation_count += 1
                cases.append({
                    "case_no": case_no,
                    "test_bau": str(test_bau),
                    "lot_line_count": lot_line_count,
                    "lot_creation_count": lot_creation_count,
                    "total_return_qty": int(total_qty) if total_qty == int(total_qty) else total_qty,
                })
                row = end_row + 1
            return cases
        finally:
            wb.close()

    def get_case_detail(self, case_no: int) -> dict[str, Any] | None:
        """Returns one unified row per non-subtotal row in the case's
        block - its A-H breakdown data (if any) and its I-M lot-creation
        fields (if any), since Created Lot# now lives on the same row as
        the original breakdown data. row_index is each row's 0-based
        position among the case's non-subtotal rows - re-derive it fresh
        on every request, it's not a stable identifier across saves.
        None if case_no doesn't exist."""
        wb = self._open_workbook()
        try:
            ws = self._get_sheet(wb)
            span = self._find_case_span(ws, case_no)
            if span is None:
                return None
            start_row, end_row = span
            test_bau = ws.cell(row=start_row, column=2).value or ""

            rows: list[dict[str, Any]] = []
            row_index = 0
            for r in range(start_row, end_row + 1):
                if self._is_subtotal_row(ws, r):
                    continue
                original_lot = ws.cell(row=r, column=3).value or ""
                created_lot_no = ws.cell(row=r, column=_LOT_CREATION_COLUMN_OFFSET + 1).value or ""
                suggested = ""
                if not created_lot_no and original_lot:
                    from app.services.lot2526.extractor import suggest_created_lot_no
                    suggested = suggest_created_lot_no(str(original_lot))
                rows.append({
                    "row_index": row_index,
                    "original_label_lot_no": str(original_lot),
                    "date_code": str(ws.cell(row=r, column=4).value or ""),
                    "return_qty_from_dc": str(ws.cell(row=r, column=5).value or ""),
                    "created_lot_no": str(created_lot_no),
                    "suggested_created_lot_no": suggested,
                    "date_created": self._format_cell_date(
                        ws.cell(row=r, column=_LOT_CREATION_COLUMN_OFFSET).value
                    ),
                    "created_date_code": str(ws.cell(row=r, column=_LOT_CREATION_COLUMN_OFFSET + 2).value or ""),
                    "physical_lot_qty": str(ws.cell(row=r, column=_LOT_CREATION_COLUMN_OFFSET + 3).value or ""),
                    "lot_code": str(ws.cell(row=r, column=_LOT_CREATION_COLUMN_OFFSET + 4).value or ""),
                })
                row_index += 1
            return {
                "case_no": case_no,
                "test_bau": str(test_bau),
                "rows": rows,
            }
        finally:
            wb.close()

    def update_row_lot_creation_fields(self, case_no: int, row_index: int, fields: dict[str, Any]) -> None:
        """Writes directly into an existing row's J/I/K/L/M cells - no row
        insertion, no shifting of anything else in the sheet. row_index is
        re-resolved fresh against the case's current non-subtotal rows
        (same ordering get_case_detail() uses), so this stays correct
        regardless of what's changed elsewhere in the sheet since the
        caller last fetched the case detail."""
        wb = self._open_workbook()
        try:
            ws = self._get_sheet(wb)
            span = self._find_case_span(ws, case_no)
            if span is None:
                raise Lot2526ExcelError(f"Case No. {case_no} not found in the 2526 worksheet")
            start_row, end_row = span

            target_row = None
            idx = 0
            for r in range(start_row, end_row + 1):
                if self._is_subtotal_row(ws, r):
                    continue
                if idx == row_index:
                    target_row = r
                    break
                idx += 1
            if target_row is None:
                raise Lot2526ExcelError(f"Row {row_index} not found in case No. {case_no}")

            self._write_lot_creation_fields(ws, target_row, fields)
            wb.save(self._masterfile_path())
        except PermissionError as exc:
            raise Lot2526ExcelLockedError(
                "Master Excel file is locked. Close it in Excel and try again."
            ) from exc
        finally:
            wb.close()

    @staticmethod
    def _write_lot_creation_fields(ws, row: int, fields: dict[str, Any]) -> None:
        date_created_raw = fields.get("date_created")
        if date_created_raw:
            try:
                date_value: Any = datetime.strptime(str(date_created_raw), "%Y-%m-%d").date()
            except ValueError:
                date_value = date_created_raw
            cell = ws.cell(row=row, column=_LOT_CREATION_COLUMN_OFFSET, value=date_value)
            cell.number_format = DATE_NUMBER_FORMAT
            cell.alignment = _CELL_ALIGNMENT

        for offset, key in enumerate(_LOT_CREATION_FIELDS[1:], start=_LOT_CREATION_COLUMN_OFFSET + 1):
            value = fields.get(key)
            if value in (None, ""):
                continue
            cell = ws.cell(row=row, column=offset, value=value)
            cell.alignment = _CELL_ALIGNMENT

    def append_lot_creation_entry(self, case_no: int, fields: dict[str, Any]) -> None:
        """Secondary action: inserts a brand-new row within an existing
        case's block carrying the 5 lot-creation fields (columns I-M), for
        when a case needs an extra row beyond its 1:1 original-lot mapping
        (e.g. a merged-wafer split into more than one created lot - not
        automated, just makes room for a manual extra entry). If the case
        ends in a subtotal row, the new row is inserted *before* it so the
        subtotal stays last; otherwise it's inserted right after the
        case's last row. Every row below (including later cases) shifts
        down by one - row positions are always re-derived fresh, never
        cached."""
        wb = self._open_workbook()
        try:
            ws = self._get_sheet(wb)
            span = self._find_case_span(ws, case_no)
            if span is None:
                raise Lot2526ExcelError(f"Case No. {case_no} not found in the 2526 worksheet")
            _, end_row = span

            insert_at = end_row if self._is_subtotal_row(ws, end_row) else end_row + 1
            ws.insert_rows(insert_at, 1)
            self._write_lot_creation_fields(ws, insert_at, fields)

            wb.save(self._masterfile_path())
        except PermissionError as exc:
            raise Lot2526ExcelLockedError(
                "Master Excel file is locked. Close it in Excel and try again."
            ) from exc
        finally:
            wb.close()

    def reset_all_data_rows(self) -> int:
        """Removes every data row (cases, continuation rows, subtotals,
        blank separators) from the 2526 sheet so the next save starts fresh
        at case No. 1. Returns the number of rows removed."""
        wb = self._open_workbook()
        try:
            ws = self._get_sheet(wb)
            last_col = 1 + len(_COLUMNS)
            removed = 0
            for row in range(ws.max_row, 1, -1):
                if any(ws.cell(row=row, column=c).value is not None for c in range(1, last_col + 1)):
                    ws.delete_rows(row, 1)
                    removed += 1
            wb.save(self._masterfile_path())
            return removed
        except PermissionError as exc:
            raise Lot2526ExcelLockedError(
                "Master Excel file is locked. Close it in Excel and try again."
            ) from exc
        finally:
            wb.close()
