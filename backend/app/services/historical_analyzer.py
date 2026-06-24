from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook

from app.services.excel_layout import WorksheetLayout, normalize_header
from config.settings import FIELD_MAPPING_PATH, resolve_masterfile_path


@dataclass
class HistoricalRow:
    sheet: str
    row_number: int
    values: dict[str, Any] = field(default_factory=dict)

    def get(self, key: str, default: Any = "") -> Any:
        val = self.values.get(key, default)
        return default if val is None else val


class HistoricalAnalyzer:
    """Loads historical rows from all FY worksheets in the master workbook."""

    def __init__(self) -> None:
        self._rows: list[HistoricalRow] = []
        self._by_dn: dict[str, HistoricalRow] = {}
        self._by_device: dict[str, list[HistoricalRow]] = {}
        self._mtime: float = 0.0
        self._field_headers = self._load_field_headers()
        self.refresh_if_stale()

    def _load_field_headers(self) -> dict[str, list[str]]:
        with open(FIELD_MAPPING_PATH, encoding="utf-8") as f:
            mapping = json.load(f)
        return mapping.get("pdf_to_excel_headers", {})

    def _reload(self) -> None:
        self._rows = []
        path = resolve_masterfile_path()
        if not path.exists():
            return

        wb = load_workbook(path, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                if not sheet_name.upper().startswith("FY"):
                    continue
                try:
                    ws = wb[sheet_name]
                    layout = WorksheetLayout.from_config_files(ws, sheet_name)
                except ValueError:
                    continue

                header_to_field: dict[str, str] = {}
                for pdf_field, aliases in self._field_headers.items():
                    for alias in aliases:
                        header_to_field[normalize_header(alias)] = pdf_field

                for row in range(layout.data_start_row, ws.max_row + 1):
                    row_data: dict[str, Any] = {}
                    has_data = False
                    for col, header in layout.headers_by_col.items():
                        value = ws.cell(row=row, column=col).value
                        if value is None:
                            continue
                        pdf_field = header_to_field.get(normalize_header(header))
                        if pdf_field:
                            row_data[pdf_field] = value
                            has_data = True
                    if has_data and (row_data.get("dn_number") or row_data.get("case_id")):
                        self._rows.append(
                            HistoricalRow(sheet=sheet_name, row_number=row, values=row_data)
                        )
        finally:
            wb.close()
        self._build_indexes()

    def _build_indexes(self) -> None:
        self._by_dn = {}
        self._by_device = {}
        for row in self._rows:
            dn = str(row.get("dn_number", "")).strip()
            if dn:
                self._by_dn[dn] = row
            device_key = self.normalize_device_key(str(row.get("device", "")))
            if device_key:
                self._by_device.setdefault(device_key, []).append(row)

    @staticmethod
    def normalize_device_key(device: str) -> str:
        return re.sub(r"\s+", "", device.upper())

    @property
    def rows(self) -> list[HistoricalRow]:
        return self._rows

    @property
    def by_dn(self) -> dict[str, HistoricalRow]:
        return self._by_dn

    @property
    def by_device(self) -> dict[str, list[HistoricalRow]]:
        return self._by_device

    def candidate_rows(self, record: dict[str, Any], limit: int = 80) -> list[HistoricalRow]:
        """Return a small set of likely historical matches instead of scanning all rows."""
        dn = str(record.get("dn_number", "")).strip()
        if dn and dn in self._by_dn:
            return [self._by_dn[dn]]

        candidates: list[HistoricalRow] = []
        seen: set[int] = set()

        def add(row: HistoricalRow) -> None:
            key = id(row)
            if key not in seen and len(candidates) < limit:
                seen.add(key)
                candidates.append(row)

        device_key = self.normalize_device_key(str(record.get("device", "")))
        if device_key:
            for row in self._by_device.get(device_key, []):
                add(row)
            if not candidates:
                for key, rows in self._by_device.items():
                    if device_key in key or key in device_key:
                        for row in rows:
                            add(row)
                        if len(candidates) >= limit:
                            break

        package = str(record.get("package", "")).strip().lower()
        if package and len(candidates) < limit:
            for row in reversed(self._rows):
                if str(row.get("package", "")).strip().lower() == package:
                    add(row)
                if len(candidates) >= limit:
                    break

        if candidates:
            return candidates

        # Small masterfiles: fall back to recent rows only
        return self._rows[-min(limit, len(self._rows)) :]

    def refresh(self) -> None:
        self._mtime = 0.0
        self.refresh_if_stale()

    def refresh_if_stale(self) -> None:
        path = resolve_masterfile_path()
        mtime = path.stat().st_mtime if path.exists() else 0.0
        if mtime == self._mtime and self._rows:
            return
        self._mtime = mtime
        self._reload()
