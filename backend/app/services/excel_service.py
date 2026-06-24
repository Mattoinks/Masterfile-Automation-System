import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.excel_layout import (
    WorksheetLayout,
    copy_row_formatting,
    load_active_worksheet_name,
    write_mapped_values,
)
from app.models.schemas import ExistingMasterRecord
from config.settings import WORKSHEET_CONFIG_PATH, resolve_masterfile_path


class ExcelServiceError(Exception):
    pass


class ExcelLockedError(ExcelServiceError):
    pass


class ExcelService:
    """
    Updates the existing master workbook in-place.
    Preserves template layout, formatting, filters, and formulas.
    Only appends or updates data rows on the configured worksheet.
    """

    def __init__(self, worksheet_name: str | None = None) -> None:
        self.worksheet_name = worksheet_name or load_active_worksheet_name()
        self._worksheet_config = self._load_worksheet_config()

    @property
    def masterfile_path(self) -> Path:
        return resolve_masterfile_path()

    def _load_worksheet_config(self) -> dict[str, Any]:
        with open(WORKSHEET_CONFIG_PATH, encoding="utf-8") as f:
            root = json.load(f)
        worksheets = root.get("worksheets", {})
        if self.worksheet_name not in worksheets:
            raise ExcelServiceError(
                f"Worksheet '{self.worksheet_name}' is not configured in worksheet_config.json"
            )
        return worksheets[self.worksheet_name]

    def ensure_masterfile_exists(self) -> None:
        if not self.masterfile_path.exists():
            raise ExcelServiceError(
                f"Master Excel template not found at {self.masterfile_path}. "
                "Place your workbook in storage/masterfile/ (RMA_MASTER.xlsx or master-template.xlsx)."
            )

    def _open_workbook(self):
        self.ensure_masterfile_exists()
        try:
            return load_workbook(self.masterfile_path)
        except PermissionError as exc:
            raise ExcelLockedError(
                "Master Excel file is locked. Close it in Excel and try again."
            ) from exc

    def _get_layout(self, wb) -> tuple[Any, WorksheetLayout]:
        if self.worksheet_name not in wb.sheetnames:
            raise ExcelServiceError(
                f"Worksheet '{self.worksheet_name}' not found in masterfile"
            )
        ws = wb[self.worksheet_name]
        layout = WorksheetLayout.from_config_files(ws, self.worksheet_name)
        return ws, layout

    def get_detected_headers(self) -> dict[str, Any]:
        wb = self._open_workbook()
        try:
            _, layout = self._get_layout(wb)
            return {
                "worksheet": self.worksheet_name,
                "headers": layout.headers_by_col,
                "pdf_field_mapping": {
                    field: layout.headers_by_col.get(col)
                    for field, col in layout.pdf_field_to_col.items()
                },
            }
        finally:
            wb.close()

    def get_all_dn_numbers(self) -> set[str]:
        wb = self._open_workbook()
        try:
            _, layout = self._get_layout(wb)
            return layout.get_all_dn_numbers()
        finally:
            wb.close()

    def get_last_case_id(self) -> int:
        wb = self._open_workbook()
        try:
            _, layout = self._get_layout(wb)
            return layout.get_max_case_id(active_only=True)
        finally:
            wb.close()

    def _prepare_insert_row(
        self, ws, layout: WorksheetLayout, record: dict[str, Any]
    ) -> tuple[int, dict[str, Any], int]:
        """Resolve target row, optionally reuse soft-deleted row for Case_ID 1."""
        target_row, case_override = layout.resolve_insert_target()
        payload = dict(record)
        if case_override is not None:
            payload["_case_id_override"] = case_override
        else:
            last_row = layout.find_last_populated_row()
            template_row = self._format_template_row(layout, last_row)
            if target_row != template_row:
                copy_row_formatting(ws, template_row, target_row)
        return target_row, payload, layout.find_last_populated_row()

    def _format_template_row(self, layout: WorksheetLayout, last_populated: int) -> int:
        if last_populated > layout.header_row:
            return last_populated
        return layout.data_start_row

    def insert_record(self, record: dict[str, Any]) -> int:
        wb = self._open_workbook()
        try:
            ws, layout = self._get_layout(wb)
            target_row, payload, _ = self._prepare_insert_row(ws, layout, record)
            values = layout.build_cell_values(payload)
            write_mapped_values(ws, target_row, values)

            case_col = layout.get_case_id_column()
            case_id = int(ws.cell(row=target_row, column=case_col).value)

            wb.save(self.masterfile_path)
            return case_id
        except PermissionError as exc:
            raise ExcelLockedError(
                "Master Excel file is locked. Close it in Excel and try again."
            ) from exc
        finally:
            wb.close()

    def replace_record(self, dn_number: str, record: dict[str, Any]) -> int:
        wb = self._open_workbook()
        try:
            ws, layout = self._get_layout(wb)
            row = layout.find_row_by_dn(dn_number)
            if row is None:
                raise ExcelServiceError(f"DN Number {dn_number} not found for replacement")

            case_col = layout.get_case_id_column()
            existing_case_id = ws.cell(row=row, column=case_col).value

            values = layout.build_cell_values(record)
            values.pop(case_col, None)
            write_mapped_values(ws, row, values)

            case_id = int(existing_case_id) if existing_case_id else layout.get_max_case_id()
            wb.save(self.masterfile_path)
            return case_id
        except PermissionError as exc:
            raise ExcelLockedError(
                "Master Excel file is locked. Close it in Excel and try again."
            ) from exc
        finally:
            wb.close()

    def get_all_master_records(self) -> list[dict[str, Any]]:
        wb = self._open_workbook()
        try:
            ws, layout = self._get_layout(wb)
            records: list[dict[str, Any]] = []
            for row_num in range(layout.data_start_row, ws.max_row + 1):
                rec = layout.row_to_record(row_num)
                if rec.get("dn_number") or rec.get("case_id"):
                    records.append(rec)
            return records
        finally:
            wb.close()

    def get_existing_record_by_dn(self, dn_number: str) -> ExistingMasterRecord | None:
        wb = self._open_workbook()
        try:
            ws, layout = self._get_layout(wb)
            row = layout.find_row_by_dn(dn_number)
            if row is None:
                return None
            rec = layout.row_to_record(row)
            return ExistingMasterRecord(
                case_id=str(rec.get("case_id", "")),
                row_number=row,
                dn_number=str(rec.get("dn_number", "")),
                rma_number=str(rec.get("rma_number", "")),
                device=str(rec.get("device", "")),
                package=str(rec.get("package", "")),
                quantity=str(rec.get("quantity", "")),
                owner=str(rec.get("owner", "")),
                date_code=str(rec.get("date_code", "")),
                store_received=str(rec.get("store_received", "")),
                case_title=str(rec.get("case_title", "")),
                type_of_return=str(rec.get("type_of_return", "")),
                gf=str(rec.get("gf", "")),
                dc=str(rec.get("dc", "")),
            )
        finally:
            wb.close()

    def search_records(
        self,
        case_id: str = "",
        dn_number: str = "",
        device: str = "",
        owner: str = "",
        package: str = "",
        sap_number: str = "",
        limit: int = 100,
    ) -> list[dict[str, Any]]:
        records = self.get_all_master_records()
        results: list[dict[str, Any]] = []

        def match(field_val: Any, query: str) -> bool:
            if not query:
                return True
            return query.strip().lower() in str(field_val or "").lower()

        for rec in records:
            if not match(rec.get("case_id"), case_id):
                continue
            if not match(rec.get("dn_number"), dn_number):
                continue
            if not match(rec.get("device"), device):
                continue
            if not match(rec.get("owner"), owner):
                continue
            if not match(rec.get("package"), package):
                continue
            if not match(rec.get("rma_number"), sap_number):
                continue
            results.append(rec)
            if len(results) >= limit:
                break
        return results

    def insert_revision(self, base_case_id: str, record: dict[str, Any]) -> int:
        wb = self._open_workbook()
        try:
            ws, layout = self._get_layout(wb)
            revision_id = layout.next_revision_case_id(base_case_id)
            record = {**record, "_case_id_override": revision_id}

            last_row = layout.find_last_populated_row()
            target_row = max(last_row + 1, layout.data_start_row)
            template_row = self._format_template_row(layout, last_row)
            if target_row != template_row:
                copy_row_formatting(ws, template_row, target_row)

            values = layout.build_cell_values(record)
            write_mapped_values(ws, target_row, values)
            wb.save(self.masterfile_path)
            return revision_id
        except PermissionError as exc:
            raise ExcelLockedError(
                "Master Excel file is locked. Close it in Excel and try again."
            ) from exc
        finally:
            wb.close()

    def force_insert_record(self, record: dict[str, Any]) -> int:
        """Insert even when DN already exists — creates new Case_ID."""
        return self.insert_record(record)

    def get_row_before_after(
        self, dn_number: str, new_record: dict[str, Any]
    ) -> dict[str, Any]:
        """Preview existing row vs proposed changes for replace."""
        existing = self.get_existing_record_by_dn(dn_number)
        if not existing:
            return {"before": None, "after": new_record}
        before = existing.model_dump()
        after = {**before, **{k: v for k, v in new_record.items() if v not in (None, "")}}
        diffs: dict[str, dict[str, str]] = {}
        for key in set(before) | set(after):
            if key in ("row_number",):
                continue
            old_v = str(before.get(key, "") or "")
            new_v = str(after.get(key, "") or "")
            if old_v != new_v:
                diffs[key] = {"existing": old_v, "new": new_v}
        return {"before": before, "after": after, "differences": diffs}

    def get_total_records(self) -> int:
        wb = self._open_workbook()
        try:
            _, layout = self._get_layout(wb)
            return len(layout.get_all_dn_numbers())
        finally:
            wb.close()

    def get_masterfile_path(self) -> Path:
        self.ensure_masterfile_exists()
        return self.masterfile_path

    def get_worksheet_rows(self, search: str = "", limit: int = 500) -> dict[str, Any]:
        wb = self._open_workbook()
        try:
            ws, layout = self._get_layout(wb)
            headers = [
                {"col": col, "label": label}
                for col, label in sorted(layout.headers_by_col.items())
            ]
            rows: list[dict[str, Any]] = []
            query = search.strip().lower()

            for row_num in range(layout.data_start_row, ws.max_row + 1):
                row_data: dict[str, Any] = {"_row": row_num}
                has_data = False
                for col, header in layout.headers_by_col.items():
                    value = ws.cell(row=row_num, column=col).value
                    if value is not None:
                        has_data = True
                    row_data[header] = value
                if not has_data:
                    continue
                if query:
                    haystack = " ".join(str(v) for v in row_data.values() if v is not None).lower()
                    if query not in haystack:
                        continue
                rows.append(row_data)
                if len(rows) >= limit:
                    break

            return {
                "worksheet": self.worksheet_name,
                "headers": headers,
                "rows": rows,
                "total": len(rows),
                "last_case_id": layout.get_max_case_id(active_only=True),
            }
        finally:
            wb.close()

    def get_worksheet_config(self) -> dict[str, Any]:
        with open(WORKSHEET_CONFIG_PATH, encoding="utf-8") as f:
            return json.load(f)

    # --- Batch session: open workbook once, save once (Steps 33–34) ---

    class _BatchSession:
        def __init__(self, service: "ExcelService") -> None:
            self.service = service
            self.wb = service._open_workbook()
            self.ws, self.layout = service._get_layout(self.wb)

        def save(self) -> None:
            try:
                self.wb.save(self.service.masterfile_path)
            except PermissionError as exc:
                raise ExcelLockedError(
                    "Master Excel file is locked. Close it in Excel and try again."
                ) from exc

        def close(self) -> None:
            self.wb.close()

        def insert_record(self, record: dict[str, Any]) -> int | str:
            target_row, payload, _ = self.service._prepare_insert_row(
                self.ws, self.layout, record
            )
            values = self.layout.build_cell_values(payload)
            write_mapped_values(self.ws, target_row, values)
            case_col = self.layout.get_case_id_column()
            val = self.ws.cell(row=target_row, column=case_col).value
            return val

        def replace_record(self, dn_number: str, record: dict[str, Any]) -> int | str:
            row = self.layout.find_row_by_dn(dn_number)
            if row is None:
                raise ExcelServiceError(f"DN Number {dn_number} not found for replacement")
            case_col = self.layout.get_case_id_column()
            existing_case_id = self.ws.cell(row=row, column=case_col).value
            values = self.layout.build_cell_values(record)
            values.pop(case_col, None)
            write_mapped_values(self.ws, row, values)
            return existing_case_id

        def insert_revision(self, base_case_id: str, record: dict[str, Any]) -> str:
            revision_id = self.layout.next_revision_case_id(base_case_id)
            record = {**record, "_case_id_override": revision_id}
            last_row = self.layout.find_last_populated_row()
            target_row = max(last_row + 1, self.layout.data_start_row)
            template_row = self.service._format_template_row(self.layout, last_row)
            if target_row != template_row:
                copy_row_formatting(self.ws, template_row, target_row)
            values = self.layout.build_cell_values(record)
            write_mapped_values(self.ws, target_row, values)
            return revision_id

        def set_record_status(self, case_id: str, status: str) -> dict[str, Any]:
            row = self.layout.find_row_by_case_id(case_id)
            if row is None:
                raise ExcelServiceError(f"Case ID {case_id} not found")
            status_col = self.layout.get_column_for_pdf_field("status")
            if status_col:
                self.ws.cell(row=row, column=status_col, value=status)
            rec = self.layout.row_to_record(row)
            rec["status"] = status
            return rec

        def update_record(self, case_id: str, fields: dict[str, Any]) -> dict[str, Any]:
            row = self.layout.find_row_by_case_id(case_id)
            if row is None:
                raise ExcelServiceError(f"Case ID {case_id} not found")
            record = self.layout.row_to_record(row)
            merged = {**record, **fields, "case_id": case_id}
            values = self.layout.build_cell_values(merged)
            case_col = self.layout.get_case_id_column()
            values.pop(case_col, None)
            write_mapped_values(self.ws, row, values)
            return self.layout.row_to_record(row)

        def permanent_delete_row(self, case_id: str) -> None:
            row = self.layout.find_row_by_case_id(case_id)
            if row is None:
                raise ExcelServiceError(f"Case ID {case_id} not found")
            self.ws.delete_rows(row, 1)

        def read_all_records(self) -> list[dict[str, Any]]:
            records: list[dict[str, Any]] = []
            for row_num in range(self.layout.data_start_row, self.ws.max_row + 1):
                rec = self.layout.row_to_record(row_num)
                if rec.get("dn_number") or rec.get("case_id"):
                    records.append(rec)
            return records

    def batch_session(self) -> _BatchSession:
        return ExcelService._BatchSession(self)

    def soft_delete_record(self, case_id: str) -> dict[str, Any]:
        session = self.batch_session()
        try:
            rec = session.set_record_status(case_id, "Deleted")
            session.save()
            return rec
        finally:
            session.close()

    def restore_record(self, case_id: str) -> dict[str, Any]:
        session = self.batch_session()
        try:
            rec = session.set_record_status(case_id, "Active")
            session.save()
            return rec
        finally:
            session.close()

    def permanent_delete_record(self, case_id: str) -> None:
        session = self.batch_session()
        try:
            session.permanent_delete_row(case_id)
            session.save()
        finally:
            session.close()

    def update_record_by_case_id(self, case_id: str, fields: dict[str, Any]) -> dict[str, Any]:
        session = self.batch_session()
        try:
            rec = session.update_record(case_id, fields)
            session.save()
            return rec
        finally:
            session.close()

    def get_record_by_case_id(self, case_id: str) -> dict[str, Any] | None:
        wb = self._open_workbook()
        try:
            _, layout = self._get_layout(wb)
            row = layout.find_row_by_case_id(case_id)
            if row is None:
                return None
            return layout.row_to_record(row)
        finally:
            wb.close()

    def sync_index_from_excel(self) -> int:
        """Full Excel → SQLite sync."""
        from app.services.index_db import get_index
        records = self.get_all_master_records()
        return get_index().sync_from_records(records, self.worksheet_name)

    def reset_all_data_rows(self) -> int:
        """Remove every data row so the next insert starts at Case_ID 1."""
        wb = self._open_workbook()
        removed = 0
        try:
            ws, layout = self._get_layout(wb)
            for row in range(ws.max_row, layout.data_start_row - 1, -1):
                if layout._row_has_data(row):
                    ws.delete_rows(row, 1)
                    removed += 1
            wb.save(self.masterfile_path)
        finally:
            wb.close()
        self.sync_index_from_excel()
        return removed
