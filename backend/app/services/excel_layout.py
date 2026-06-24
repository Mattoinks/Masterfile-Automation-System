from __future__ import annotations

import json
import re
from copy import copy
from pathlib import Path
from typing import Any

from app.services.status_utils import is_active_status, normalize_record_status

from openpyxl.worksheet.worksheet import Worksheet

from config.settings import FIELD_MAPPING_PATH, WORKSHEET_CONFIG_PATH


def normalize_header(text: str) -> str:
    if text is None:
        return ""
    cleaned = str(text).replace("\n", " ").replace("\r", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip().lower()
    return cleaned


class WorksheetLayout:
    """Reads column headers from the workbook and maps PDF fields to columns."""

    def __init__(
        self,
        ws: Worksheet,
        worksheet_name: str,
        worksheet_config: dict[str, Any],
        field_mapping: dict[str, Any],
    ) -> None:
        self.ws = ws
        self.worksheet_name = worksheet_name
        self.worksheet_config = worksheet_config
        self.field_mapping = field_mapping
        self.header_row = worksheet_config.get("header_row", 1)
        self.data_start_row = worksheet_config.get("data_start_row", 2)
        self.default_fy = worksheet_config.get("default_fy", worksheet_name)
        self.headers_by_col: dict[int, str] = {}
        self.col_by_normalized_header: dict[str, int] = {}
        self.pdf_field_to_col: dict[str, int] = {}
        self._scan_headers()
        self._build_pdf_mapping()

    @classmethod
    def from_config_files(cls, ws: Worksheet, worksheet_name: str) -> WorksheetLayout:
        with open(WORKSHEET_CONFIG_PATH, encoding="utf-8") as f:
            ws_config_root = json.load(f)
        with open(FIELD_MAPPING_PATH, encoding="utf-8") as f:
            field_mapping = json.load(f)

        worksheets = ws_config_root.get("worksheets", {})
        if worksheet_name not in worksheets:
            raise ValueError(f"Worksheet '{worksheet_name}' is not configured")

        return cls(ws, worksheet_name, worksheets[worksheet_name], field_mapping)

    def _scan_headers(self) -> None:
        for col in range(1, self.ws.max_column + 1):
            value = self.ws.cell(row=self.header_row, column=col).value
            if value is None:
                continue
            header = str(value).strip()
            self.headers_by_col[col] = header
            self.col_by_normalized_header[normalize_header(header)] = col

    def _build_pdf_mapping(self) -> None:
        header_aliases: dict[str, list[str]] = self.field_mapping.get("pdf_to_excel_headers", {})
        for pdf_field, aliases in header_aliases.items():
            for alias in aliases:
                col = self.col_by_normalized_header.get(normalize_header(alias))
                if col is not None:
                    self.pdf_field_to_col[pdf_field] = col
                    break

    def get_column_for_pdf_field(self, pdf_field: str) -> int | None:
        return self.pdf_field_to_col.get(pdf_field)

    def get_column_for_header(self, header: str) -> int | None:
        return self.col_by_normalized_header.get(normalize_header(header))

    def get_dn_column(self) -> int:
        col = self.get_column_for_pdf_field("dn_number")
        if col is None:
            raise ValueError("DN / Invoice No. column not found in worksheet headers")
        return col

    def get_case_id_column(self) -> int:
        col = self.get_column_for_pdf_field("case_id")
        if col is None:
            raise ValueError("Case_ID column not found in worksheet headers")
        return col

    def find_last_populated_row(self) -> int:
        """Last row with Case_ID or DN number populated."""
        case_col = self.get_case_id_column()
        dn_col = self.get_dn_column()
        last_row = self.header_row
        for row in range(self.data_start_row, self.ws.max_row + 1):
            case_val = self.ws.cell(row=row, column=case_col).value
            dn_val = self.ws.cell(row=row, column=dn_col).value
            if case_val is not None or dn_val is not None:
                last_row = row
        return last_row

    def find_row_by_dn(self, dn_number: str) -> int | None:
        dn_col = self.get_dn_column()
        target = str(dn_number).strip()
        for row in range(self.data_start_row, self.ws.max_row + 1):
            value = self.ws.cell(row=row, column=dn_col).value
            if value is not None and str(value).strip() == target:
                return row
        return None

    def get_all_dn_numbers(self) -> set[str]:
        dn_col = self.get_dn_column()
        numbers: set[str] = set()
        for row in range(self.data_start_row, self.ws.max_row + 1):
            value = self.ws.cell(row=row, column=dn_col).value
            if value is not None:
                numbers.add(str(value).strip())
        return numbers

    def _row_has_data(self, row_num: int) -> bool:
        case_col = self.get_case_id_column()
        dn_col = self.get_dn_column()
        case_val = self.ws.cell(row=row_num, column=case_col).value
        dn_val = self.ws.cell(row=row_num, column=dn_col).value
        return case_val is not None or dn_val is not None

    def _row_status(self, row_num: int) -> str:
        status_col = self.get_column_for_pdf_field("status")
        if not status_col:
            return "Open"
        return normalize_record_status(self.ws.cell(row=row_num, column=status_col).value)

    def is_row_deleted(self, row_num: int) -> bool:
        if not self._row_has_data(row_num):
            return False
        return not is_active_status(self._row_status(row_num))

    def count_active_rows(self) -> int:
        count = 0
        for row in range(self.data_start_row, self.ws.max_row + 1):
            if self._row_has_data(row) and not self.is_row_deleted(row):
                count += 1
        return count

    def find_first_deleted_row(self) -> int | None:
        for row in range(self.data_start_row, self.ws.max_row + 1):
            if self.is_row_deleted(row):
                return row
        return None

    def get_max_case_id(self, active_only: bool = False) -> int:
        case_col = self.get_case_id_column()
        max_id = 0
        for row in range(self.data_start_row, self.ws.max_row + 1):
            if not self._row_has_data(row):
                continue
            if active_only and self.is_row_deleted(row):
                continue
            value = self.ws.cell(row=row, column=case_col).value
            if value is not None:
                try:
                    max_id = max(max_id, int(float(str(value))))
                except (ValueError, TypeError):
                    continue
        return max_id

    def get_next_case_id(self) -> int:
        active_max = self.get_max_case_id(active_only=True)
        if active_max > 0:
            return active_max + 1
        return 1

    def resolve_insert_target(self) -> tuple[int, int | None]:
        """Pick row for insert; reuse a soft-deleted row when no active records remain."""
        if self.count_active_rows() == 0:
            deleted_row = self.find_first_deleted_row()
            if deleted_row is not None:
                rec = self.row_to_record(deleted_row)
                try:
                    case_id = int(float(str(rec.get("case_id", 1))))
                except (ValueError, TypeError):
                    case_id = 1
                return deleted_row, case_id

        last_row = self.find_last_populated_row()
        target_row = max(last_row + 1, self.data_start_row)
        return target_row, None

    def get_all_case_ids(self) -> set[str]:
        case_col = self.get_case_id_column()
        ids: set[str] = set()
        for row in range(self.data_start_row, self.ws.max_row + 1):
            value = self.ws.cell(row=row, column=case_col).value
            if value is not None:
                ids.add(str(value).strip())
        return ids

    def row_to_record(self, row_num: int) -> dict[str, Any]:
        """Read a worksheet row into PDF field names."""
        col_to_field: dict[int, str] = {}
        for field, col in self.pdf_field_to_col.items():
            col_to_field[col] = field

        record: dict[str, Any] = {"row_number": row_num}
        for col, field in col_to_field.items():
            value = self.ws.cell(row=row_num, column=col).value
            if value is not None:
                record[field] = value
        if "case_id" in record:
            record["case_id"] = str(record["case_id"]).strip()
        if "status" in record:
            record["status"] = normalize_record_status(record["status"])
        return record

    def find_row_by_case_id(self, case_id: str) -> int | None:
        case_col = self.get_case_id_column()
        target = str(case_id).strip()
        for row in range(self.data_start_row, self.ws.max_row + 1):
            value = self.ws.cell(row=row, column=case_col).value
            if value is not None and str(value).strip() == target:
                return row
        return None

    def next_revision_case_id(self, base_case_id: str) -> str:
        """Generate revision suffix: 1866 -> 1866-R1, or 1866A if -R pattern unused."""
        existing = self.get_all_case_ids()
        base = str(base_case_id).strip()
        revision = 1
        while f"{base}-R{revision}" in existing:
            revision += 1
        candidate = f"{base}-R{revision}"
        if candidate not in existing:
            return candidate
        for suffix in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
            alt = f"{base}{suffix}"
            if alt not in existing:
                return alt
        return f"{base}-R{revision + 1}"

    def build_cell_values(self, record: dict[str, Any]) -> dict[int, Any]:
        """Map extracted PDF record to column indexes."""
        computed = self.field_mapping.get("computed_fields", {})
        defaults = self.field_mapping.get("config_defaults", {})

        values: dict[int, Any] = {}

        if computed.get("case_id") == "auto_increment":
            col = self.get_column_for_pdf_field("case_id")
            if col:
                override = record.get("_case_id_override")
                if override:
                    values[col] = override
                else:
                    values[col] = self.get_next_case_id()

        field_sources = {
            "fy": self.default_fy,
            "data_source": record.get("data_source") or defaults.get("data_source", "SAP DN"),
            "status": defaults.get("status", "Open"),
            "dn_number": record.get("dn_number", ""),
            "dn_date": record.get("dn_date", ""),
            "store_received": record.get("store_received") or record.get("dn_date", ""),
            "recd_lw": record.get("recd_lw", ""),
            "recd_mth": record.get("recd_mth", ""),
            "rma_number": record.get("rma_number", ""),
            "type_of_return": record.get("type_of_return", ""),
            "device": record.get("device", ""),
            "package": record.get("package", ""),
            "gf": record.get("gf", ""),
            "date_code": record.get("date_code", ""),
            "dc_bau": record.get("dc_bau", ""),
            "test_bau": record.get("test_bau", ""),
            "vkl_bau": record.get("vkl_bau", ""),
            "quantity": self._coerce_quantity(record.get("quantity", "")),
            "dc": record.get("dc", ""),
            "owner": record.get("owner", ""),
            "rework_flow_procedure": record.get("rework_flow_procedure", ""),
            "cause_owner": record.get("cause_owner", ""),
            "case_title": record.get("case_title", ""),
            "store_lot_qty": record.get("store_lot_qty", ""),
            "status": normalize_record_status(record.get("status") or defaults.get("status", "Open")),
        }

        for field, value in field_sources.items():
            if value in ("", None):
                continue
            col = self.get_column_for_pdf_field(field)
            if col is not None and col not in values:
                values[col] = value

        return values

    @staticmethod
    def _coerce_quantity(raw: Any) -> int | str:
        if raw in ("", None):
            return ""
        try:
            return int(float(str(raw).replace(",", "")))
        except (ValueError, TypeError):
            return raw


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.number_format = copy(source.number_format)
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)


def adjust_formula_row(formula: str, source_row: int, target_row: int) -> str:
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    return re.sub(rf"([A-Z]{{1,3}}){source_row}(?![0-9])", rf"\1{target_row}", formula)


def copy_row_formatting(ws: Worksheet, source_row: int, target_row: int) -> None:
    """Copy row height, cell formatting, and row-relative formulas from source to target."""
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    max_col = max(ws.max_column, 1)
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        tgt = ws.cell(row=target_row, column=col)
        copy_cell_style(src, tgt)

        if isinstance(src.value, str) and src.value.startswith("="):
            tgt.value = adjust_formula_row(src.value, source_row, target_row)
        else:
            tgt.value = None


def write_mapped_values(ws: Worksheet, target_row: int, values: dict[int, Any]) -> None:
    for col, value in values.items():
        if value not in ("", None):
            ws.cell(row=target_row, column=col, value=value)


def load_active_worksheet_name() -> str:
    with open(WORKSHEET_CONFIG_PATH, encoding="utf-8") as f:
        config = json.load(f)
    return config.get("active_worksheet", "FY2526")
